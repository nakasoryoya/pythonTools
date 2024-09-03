import openpyxl as xl
import datetime

import myutil.configReader as configReader
import myutil.dbutil as dbutil
import myutil.excelutil as excelutil


def main():
    conf = configReader.read_configuration('setting.conf', 'connection')

    database_accessor = dbutil.DatabaseAccessor(
        conf['driver'], conf['server'], conf['database'], conf['username'], conf['password'])

    # カラム名を取得
    databases = conf['target_database'].split(',')

    wb = xl.Workbook()
    # データベースごとのサイズを取得し、テキストファイルに出力
    database_sizes = []
    for db in databases:
        try:
            database_accessor.execute_update_sql("use " + db)
        except:
            continue

        database_sizes.append(database_accessor.execute_sql("EXEC sp_spaceused")[0])

        tables = database_accessor.execute_sql("select name from sys.tables")
        table_sizes = []
        for table in tables:
            try:
                table_sizes.append(database_accessor.execute_sql("EXEC sp_spaceused " + table['name'])[0])
            except:
                continue

        # サイズの降順でソート
        table_sizes.sort(key=lambda x: int(x['reserved'][:-3]), reverse=True)

        ws = wb.create_sheet(db)
        excelutil.output_dictionary_to_excel(table_sizes, ws)

    database_sizes.sort(key=lambda x: float(x['database_size'][:-3]), reverse=True)
    # ファイル出力
    ws = wb.create_sheet('databases', 0)
    excelutil.output_dictionary_to_excel(database_sizes, ws)
    wb.remove(wb['Sheet'])

    wb.save(f"databases_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx")


if __name__ == "__main__":
    main()
