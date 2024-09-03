import openpyxl as xl
import win32com.client as client

import warnings

# DrawingMLに関する警告を無視する
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


class WorkBook:
    def __init__(self, path=None):
        if path:
            self.wb = xl.load_workbook(path)
        else:
            self.wb = xl.Workbook()

        self.path = path
        self.name = self.path.split("\\")[-1]

    def get_worksheet(self, sheet_name):
        for ws in self.get_ws_name_list():
            if ws.get_name() == sheet_name:
                return ws
        return None

    def create_worksheet(self, sheet_name):
        ws = WorkSheet(self.wb, sheet_name)
        return ws

    def delete_worksheet(self, sheet_name):
        ws = self.get_worksheet(sheet_name)
        if ws:
            self.wb.remove(ws.ws)

    def save(self, path=None):
        if path:
            self.wb.save(path)
        else:
            self.wb.save(self.path)

    def get_ws_name_list(self):
        return [WorkSheet(self.wb, sheet_name) for sheet_name in self.wb.sheetnames]

    def close(self, is_force=False):
        target_wb = client.Dispatch("Excel.Application").Workbooks(self.name)
        target_wb.Close(is_force)

    def is_saved(self):
        return client.Dispatch("Excel.Application").Workbooks(self.name).Saved


class WorkSheet:
    def __init__(self, wb, sheet_name):
        # シートが存在しない場合は新規作成、存在する場合は取得
        if sheet_name in wb.sheetnames:
            self.ws = wb[sheet_name]
        else:
            self.ws = wb.create_sheet(sheet_name)

        self.sheet_name = sheet_name

    def append(self, row):
        self.ws.append(row)

    def output_dictionary_to_excel(self, output_dict, start_row=1, start_col=1):
        self.ws.append(list(output_dict[0].keys()))
        for row in output_dict:
            self.ws.append(list(row.values()))

    def output_list_to_excel(self, output_list, start_row=1, start_col=1):
        for row in output_list:
            self.ws.append([row])

    def input_range_into_list(self, start_row, end_row, start_col, end_col):
        row_list = []
        for row in self.ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
            row_list.append([cell.value for cell in row])
        return row_list

    def get_name(self):
        return self.sheet_name


def get_open_book_list():
    excel = client.Dispatch("Excel.Application")
    workbook_list = []
    for i in range(1, excel.Workbooks.Count + 1):
        workbook = excel.Workbooks(i)
        workbook_list.append(WorkBook(workbook.FullName))
    return workbook_list
